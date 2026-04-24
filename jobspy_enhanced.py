#!/usr/bin/env python
"""Enhanced JobSpy scraper for maximum fresher SDE job discovery."""

from __future__ import annotations

import argparse
import csv
import json
import logging
import random
import re
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime
from itertools import cycle, product
from pathlib import Path
from typing import Any, Callable, Iterable

import pandas as pd
import yaml
from fake_useragent import UserAgent
from openpyxl.styles import Font, PatternFill
from rapidfuzz import fuzz
from tqdm import tqdm

from jobspy import scrape_jobs


DEFAULT_CONFIG_PATH = Path("config.yaml")
DEFAULT_SEEN_JOBS_PATH = Path("seen_jobs.csv")

ALLOWED_PLATFORMS = ["linkedin", "indeed", "glassdoor", "zip_recruiter", "google"]
TYPE_ORDER = ["A", "B", "C", "D", "E", "F", "G"]

TITLE_STRICT_PATTERN = re.compile(
    r"\b(senior|lead|principal|staff|director|vp|head of|architect|manager|president)\b",
    re.IGNORECASE,
)
HIGH_EXPERIENCE_PATTERN = re.compile(
    r"\b(5\+\s*years?|7\+\s*years?|10\+\s*years?|minimum\s+5\s+years?)\b",
    re.IGNORECASE,
)

TITLE_FRESHER_SIGNAL_PATTERN = re.compile(
    r"\b(fresher|junior|entry|associate|trainee|graduate)\b",
    re.IGNORECASE,
)
DESC_POSITIVE_EXPERIENCE_PATTERN = re.compile(
    r"\b(0\s*[-to]\s*2\s*years?|0\s*[-to]\s*1\s*years?|no\s+experience)\b",
    re.IGNORECASE,
)
DESC_THREE_PLUS_PATTERN = re.compile(r"\b3\+\s*years?\b", re.IGNORECASE)
DESC_FIVE_PLUS_PATTERN = re.compile(r"\b5\+\s*years?\b", re.IGNORECASE)
CORE_CS_PATTERN = re.compile(
    r"\b(oop|oops|dsa|data\s+structures?|algorithms?|dbms|operating\s+systems?|computer\s+networks?|cn|os)\b",
    re.IGNORECASE,
)
EXPERIENCE_MENTION_PATTERN = re.compile(
    r"\b(\d+\+?\s*years?|\d+\s*[-to]\s*\d+\s*years?|no\s+experience|"
    r"fresher|entry\s*level|new\s+grad|recent\s+graduate|campus\s+hire)\b",
    re.IGNORECASE,
)

BOND_TERMS = [
    "bond",
    "service agreement",
    "training bond",
    "2 year bond",
    "pay back",
    "clawback",
    "penalty clause",
]

OUTPUT_COLUMNS = [
    "title",
    "company",
    "location",
    "date_posted",
    "job_type",
    "relevance_score",
    "skills_matched",
    "experience_mentioned",
    "min_amount",
    "max_amount",
    "salary_currency",
    "job_url",
    "found_on_platforms",
    "multi_platform_hit",
    "bond_flag",
    "bond_details",
    "description_full",
    "description_snippet",
    "combo_that_found_it",
    "scrape_timestamp",
]

COLOR_DARK_GREEN = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
COLOR_LIGHT_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
COLOR_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
COLOR_ORANGE = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
COLOR_RED = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
BOND_FONT = Font(color="FF0000")


@dataclass
class ComboQuery:
    """Represents a generated search-term combination."""

    combo_id: str
    combo_type: str
    query: str


@dataclass
class ComboTask:
    """Represents a combo and the locations it should run against."""

    order: int
    combo: ComboQuery
    locations_to_search: list[str]


@dataclass
class ComboResult:
    """Stores per-combo execution output and metrics."""

    task: ComboTask
    jobs_df: pd.DataFrame
    platform_counts: dict[str, int]
    skipped_platforms: list[str]


def parse_bool(value: str) -> bool:
    """Parses a CLI boolean string safely."""
    val = str(value).strip().lower()
    if val in {"1", "true", "t", "yes", "y"}:
        return True
    if val in {"0", "false", "f", "no", "n"}:
        return False
    raise argparse.ArgumentTypeError(f"Invalid boolean value: {value}")


def parse_csv_list(value: str | None) -> list[str]:
    """Parses comma-separated CLI strings into a clean list."""
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def setup_logging(timestamp: str) -> tuple[logging.Logger, Path]:
    """Configures console + file logging for one scraper run."""
    log_path = Path(f"scraper_log_{timestamp}.txt")
    logger = logging.getLogger("jobspy_enhanced")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    return logger, log_path


def load_config(config_path: Path) -> dict[str, Any]:
    """Loads YAML config banks and defaults."""
    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")
    with config_path.open("r", encoding="utf-8") as handle:
        config = yaml.safe_load(handle) or {}
    if not isinstance(config, dict):
        raise ValueError("config.yaml must contain a top-level mapping")
    return config


def normalize_platform_name(platform: str) -> str:
    """Normalizes platform names from user input to JobSpy-supported values."""
    p = platform.strip().lower().replace("-", "_")
    if p == "ziprecruiter":
        return "zip_recruiter"
    return p


def resolve_platforms(config: dict[str, Any], cli_platforms: str | None) -> list[str]:
    """Resolves requested platforms with validation."""
    default_platforms = config.get("platforms", ALLOWED_PLATFORMS)
    requested = parse_csv_list(cli_platforms) if cli_platforms else list(default_platforms)
    normalized = [normalize_platform_name(item) for item in requested]
    valid = [item for item in normalized if item in ALLOWED_PLATFORMS]
    if not valid:
        raise ValueError(
            f"No valid platforms selected. Allowed: {', '.join(ALLOWED_PLATFORMS)}"
        )
    return list(dict.fromkeys(valid))


def load_proxies(proxy_path: str | None) -> list[str] | None:
    """Loads optional proxies from a text file."""
    if not proxy_path:
        return None
    path = Path(proxy_path)
    if not path.exists():
        raise FileNotFoundError(f"Proxy file not found: {proxy_path}")
    proxies: list[str] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            value = line.strip()
            if value and not value.startswith("#"):
                proxies.append(value)
    return proxies or None


def resolve_bank(config: dict[str, Any], key: str, cli_value: str | None) -> list[str]:
    """Resolves a keyword bank from CLI override or config defaults."""
    cli_items = parse_csv_list(cli_value)
    if cli_items:
        return cli_items
    items = config.get(key, [])
    if not isinstance(items, list) or not items:
        raise ValueError(f"Config key '{key}' must be a non-empty list")
    return [str(item).strip() for item in items if str(item).strip()]


def compute_type_targets(min_combos: int) -> dict[str, int]:
    """Computes balanced per-type combo targets that total at least min_combos."""
    ratios = {
        "A": 0.22,
        "B": 0.20,
        "C": 0.14,
        "D": 0.14,
        "E": 0.10,
        "F": 0.08,
        "G": 0.12,
    }
    targets = {combo_type: max(1, int(min_combos * ratio)) for combo_type, ratio in ratios.items()}
    while sum(targets.values()) < min_combos:
        for combo_type in TYPE_ORDER:
            targets[combo_type] += 1
            if sum(targets.values()) >= min_combos:
                break
    return targets


def next_unique_query(
    iterator: Iterable[str],
    seen_queries: set[str],
) -> str | None:
    """Pulls the next unique query string from a generator-like iterable."""
    for query in iterator:
        cleaned = " ".join(query.split()).strip()
        if cleaned and cleaned not in seen_queries:
            seen_queries.add(cleaned)
            return cleaned
    return None


def generate_search_combinations(
    job_titles: list[str],
    skills: list[str],
    experience_qualifiers: list[str],
    locations: list[str],
    min_combos: int,
    seed: int = 42,
) -> list[ComboQuery]:
    """Generates unique search combinations across Types A-G using itertools.product."""
    targets = compute_type_targets(max(100, min_combos))
    seen_queries: set[str] = set()

    type_generators: dict[str, Iterable[str]] = {
        "A": (
            f"{title} {skill} {exp}"
            for title, skill, exp in product(job_titles, skills, experience_qualifiers)
        ),
        "B": (
            f"{title} {skill} {location}"
            for title, skill, location in product(job_titles, skills, locations)
        ),
        "C": (f"{title} {exp}" for title, exp in product(job_titles, experience_qualifiers)),
        "D": (
            f"{skill} {exp} {location}"
            for skill, exp, location in product(skills, experience_qualifiers, locations)
        ),
        "E": (f"{title}" for (title,) in product(job_titles)),
        "F": (
            f"{exp} {location}" for exp, location in product(experience_qualifiers, locations)
        ),
        "G": (f"{skill} {location}" for skill, location in product(skills, locations)),
    }

    iterators = {combo_type: iter(generator) for combo_type, generator in type_generators.items()}
    staged_combos: list[tuple[str, str]] = []
    counts = Counter()

    for combo_type in TYPE_ORDER:
        target = targets[combo_type]
        while counts[combo_type] < target:
            query = next_unique_query(iterators[combo_type], seen_queries)
            if query is None:
                break
            staged_combos.append((combo_type, query))
            counts[combo_type] += 1

    while len(staged_combos) < max(100, min_combos):
        progressed = False
        for combo_type in TYPE_ORDER:
            query = next_unique_query(iterators[combo_type], seen_queries)
            if query is None:
                continue
            staged_combos.append((combo_type, query))
            counts[combo_type] += 1
            progressed = True
            if len(staged_combos) >= max(100, min_combos):
                break
        if not progressed:
            break

    rng = random.Random(seed)
    rng.shuffle(staged_combos)

    combos: list[ComboQuery] = []
    for idx, (combo_type, query) in enumerate(staged_combos, start=1):
        combos.append(ComboQuery(combo_id=f"C{idx:03d}", combo_type=combo_type, query=query))
    return combos


def build_combo_tasks(combos: list[ComboQuery], locations: list[str]) -> list[ComboTask]:
    """Builds combo tasks with required India/Remote coverage and city rotation."""
    city_locations = [
        loc
        for loc in locations
        if loc.strip().lower() not in {"india", "remote"}
    ]
    city_cycle = cycle(city_locations if city_locations else ["Bangalore"])

    tasks: list[ComboTask] = []
    for order, combo in enumerate(combos, start=1):
        target_locations = ["India", "Remote"]
        if combo.combo_type in {"B", "G"}:
            target_locations.append(next(city_cycle))
        deduped_locations = list(dict.fromkeys(loc.strip() for loc in target_locations if loc.strip()))
        tasks.append(
            ComboTask(
                order=order,
                combo=combo,
                locations_to_search=deduped_locations,
            )
        )
    return tasks


def canonicalize_url(value: Any) -> str:
    """Normalizes job URLs for deduplication and history tracking."""
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return ""
    return text.rstrip("/").lower()


def normalize_text(value: Any) -> str:
    """Normalizes text for robust matching and deduplication."""
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def first_non_empty(series: pd.Series) -> Any:
    """Returns the first non-empty value from a series."""
    for value in series:
        if pd.notna(value) and str(value).strip() and str(value).strip().lower() != "nan":
            return value
    return None


def longest_text(series: pd.Series) -> str:
    """Returns the longest non-empty text value from a series."""
    best = ""
    for value in series:
        text = str(value or "")
        if len(text) > len(best):
            best = text
    return best


def unique_join(values: pd.Series, split_pattern: str, separator: str) -> str:
    """Returns ordered unique values from a tokenized string series."""
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if pd.isna(value):
            continue
        for part in re.split(split_pattern, str(value)):
            item = part.strip()
            if item and item not in seen:
                seen.add(item)
                ordered.append(item)
    return separator.join(ordered)


def ensure_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Ensures a DataFrame has all required columns."""
    for column in columns:
        if column not in df.columns:
            df[column] = None
    return df


def initialize_user_agent_provider(logger: logging.Logger) -> Callable[[], str | None]:
    """Initializes fake_useragent with a safe fallback provider."""
    fallback_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    ]

    try:
        provider = UserAgent()

        def getter() -> str | None:
            """Returns a rotating user-agent string."""
            try:
                return provider.random
            except Exception:
                return random.choice(fallback_agents)

        return getter
    except Exception as exc:
        logger.warning("fake_useragent unavailable (%s). Using fallback user-agents.", exc)

        def fallback_getter() -> str | None:
            """Returns a fallback user-agent string."""
            return random.choice(fallback_agents)

        return fallback_getter


def scrape_location_with_retry(
    *,
    search_term: str,
    location: str,
    platforms: list[str],
    results_wanted: int,
    hours_old: int,
    country_indeed: str,
    proxies: list[str] | None,
    user_agent_getter: Callable[[], str | None],
    retry_backoff_seconds: list[int],
    logger: logging.Logger,
) -> tuple[pd.DataFrame, dict[str, int], list[str]]:
    """Scrapes a single location with retries and platform-level skip tracking."""
    platform_fail_counts: Counter[str] = Counter()
    skipped_platforms: list[str] = []
    active_platforms = list(platforms)
    last_frame = pd.DataFrame()
    last_metadata: dict[str, Any] = {}

    for attempt in range(1, 4):
        if not active_platforms:
            break

        try:
            frame = scrape_jobs(
                site_name=active_platforms,
                search_term=search_term,
                location=location,
                results_wanted=results_wanted,
                hours_old=hours_old,
                country_indeed=country_indeed,
                linkedin_fetch_description=True,
                proxies=proxies,
                user_agent=user_agent_getter(),
                verbose=0,
            )
        except Exception as exc:
            logger.warning(
                "Location scrape failed | '%s' | location=%s | attempt %s/3 | %s",
                search_term,
                location,
                attempt,
                exc,
            )
            frame = pd.DataFrame()
            last_metadata = {platform: {"status": "error", "returned": 0} for platform in active_platforms}
        else:
            last_metadata = getattr(frame, "attrs", {}).get("scrape_metadata", {}) or {}
            last_frame = frame if isinstance(frame, pd.DataFrame) else pd.DataFrame()

        if not last_metadata:
            last_metadata = {platform: {"status": "ok", "returned": 0} for platform in active_platforms}

        statuses = []
        for platform in list(active_platforms):
            status = str(last_metadata.get(platform, {}).get("status", "ok")).lower()
            statuses.append(status)
            if status == "error":
                platform_fail_counts[platform] += 1
                if platform_fail_counts[platform] >= 3:
                    active_platforms.remove(platform)
                    if platform not in skipped_platforms:
                        skipped_platforms.append(platform)
                        logger.warning(
                            "Platform skipped after 3 failed attempts | platform=%s | query='%s'",
                            platform,
                            search_term,
                        )
            else:
                platform_fail_counts[platform] = 0

        if any(status != "error" for status in statuses):
            platform_counts = {
                platform: int(last_metadata.get(platform, {}).get("returned", 0) or 0)
                for platform in platforms
            }
            return last_frame, platform_counts, skipped_platforms

        if attempt < 3:
            sleep_for = retry_backoff_seconds[attempt - 1]
            logger.info(
                "Retrying failed location scrape in %ss | '%s' | location=%s | attempt=%s",
                sleep_for,
                search_term,
                location,
                attempt + 1,
            )
            time.sleep(sleep_for)

    platform_counts = {
        platform: int(last_metadata.get(platform, {}).get("returned", 0) or 0)
        for platform in platforms
    }
    return last_frame, platform_counts, skipped_platforms


def run_combo_task(
    task: ComboTask,
    *,
    platforms: list[str],
    results_wanted: int,
    hours_old: int,
    country_indeed: str,
    proxies: list[str] | None,
    user_agent_getter: Callable[[], str | None],
    retry_backoff_seconds: list[int],
    sleep_range_seconds: tuple[float, float],
    scrape_timestamp: str,
    logger: logging.Logger,
) -> ComboResult:
    """Runs all location scrapes for one combo task."""
    frames: list[pd.DataFrame] = []
    platform_totals: Counter[str] = Counter()
    skipped_platforms: list[str] = []

    for location in task.locations_to_search:
        time.sleep(random.uniform(sleep_range_seconds[0], sleep_range_seconds[1]))
        frame, platform_counts, skipped = scrape_location_with_retry(
            search_term=task.combo.query,
            location=location,
            platforms=platforms,
            results_wanted=results_wanted,
            hours_old=hours_old,
            country_indeed=country_indeed,
            proxies=proxies,
            user_agent_getter=user_agent_getter,
            retry_backoff_seconds=retry_backoff_seconds,
            logger=logger,
        )
        platform_totals.update(platform_counts)
        for platform in skipped:
            if platform not in skipped_platforms:
                skipped_platforms.append(platform)

        if isinstance(frame, pd.DataFrame) and not frame.empty:
            enriched = frame.copy()
            enriched["combo_id"] = task.combo.combo_id
            enriched["combo_type"] = task.combo.combo_type
            enriched["combo_that_found_it"] = task.combo.query
            enriched["search_location"] = location
            enriched["scrape_timestamp"] = scrape_timestamp
            frames.append(enriched)

    if frames:
        jobs_df = pd.concat(frames, ignore_index=True)
    else:
        jobs_df = pd.DataFrame()

    return ComboResult(
        task=task,
        jobs_df=jobs_df,
        platform_counts={platform: int(platform_totals.get(platform, 0)) for platform in platforms},
        skipped_platforms=skipped_platforms,
    )


def save_checkpoint(raw_frames: list[pd.DataFrame], timestamp: str, completed_combos: int) -> Path | None:
    """Saves periodic checkpoint CSV so partial progress is never lost."""
    if not raw_frames:
        return None
    checkpoint_dir = Path("checkpoints")
    checkpoint_dir.mkdir(parents=True, exist_ok=True)
    checkpoint_path = checkpoint_dir / f"partial_results_{timestamp}_combo_{completed_combos}.csv"
    checkpoint_df = pd.concat(raw_frames, ignore_index=True)
    checkpoint_df.to_csv(
        checkpoint_path,
        index=False,
        quoting=csv.QUOTE_NONNUMERIC,
        escapechar="\\",
    )
    return checkpoint_path


def standardize_raw_jobs(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Normalizes raw JobSpy schema into the reporting schema."""
    if raw_df.empty:
        return raw_df

    required = [
        "title",
        "company",
        "location",
        "date_posted",
        "job_type",
        "min_amount",
        "max_amount",
        "currency",
        "job_url",
        "job_url_direct",
        "description",
        "site",
        "is_remote",
        "combo_that_found_it",
        "scrape_timestamp",
        "search_location",
        "combo_type",
        "combo_id",
    ]
    df = ensure_columns(raw_df.copy(), required)

    df["job_url"] = df["job_url"].where(
        df["job_url"].notna() & (df["job_url"].astype(str).str.strip() != ""),
        df["job_url_direct"],
    )
    df["salary_currency"] = df["currency"]
    df["description_full"] = df["description"].fillna("").astype(str)
    df["description_snippet"] = df["description_full"].str.slice(0, 500)
    df["found_on_platforms"] = df["site"].fillna("").astype(str).str.lower()
    df["multi_platform_hit"] = False
    df["date_posted"] = pd.to_datetime(df["date_posted"], errors="coerce").dt.date

    return df[
        [
            "title",
            "company",
            "location",
            "date_posted",
            "job_type",
            "min_amount",
            "max_amount",
            "salary_currency",
            "job_url",
            "description_full",
            "description_snippet",
            "found_on_platforms",
            "multi_platform_hit",
            "combo_that_found_it",
            "scrape_timestamp",
            "is_remote",
            "search_location",
            "combo_type",
            "combo_id",
        ]
    ]


def apply_fresher_filter(df: pd.DataFrame) -> pd.DataFrame:
    """Filters out clearly non-fresher jobs only when both strict rules match."""
    if df.empty:
        return df

    title_series = df["title"].fillna("").astype(str)
    desc_series = df["description_full"].fillna("").astype(str)

    senior_title_mask = title_series.str.contains(TITLE_STRICT_PATTERN)
    high_exp_desc_mask = desc_series.str.contains(HIGH_EXPERIENCE_PATTERN)

    return df.loc[~(senior_title_mask & high_exp_desc_mask)].copy()


def add_normalized_keys(df: pd.DataFrame) -> pd.DataFrame:
    """Adds temporary normalized columns used for deduplication."""
    working = df.copy()
    working["company_norm"] = working["company"].apply(normalize_text)
    working["title_norm"] = working["title"].apply(normalize_text)
    working["location_norm"] = working["location"].apply(normalize_text)
    working["date_norm"] = pd.to_datetime(working["date_posted"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
    working["job_url_norm"] = working["job_url"].apply(canonicalize_url)
    working["company_title_norm"] = (working["company_norm"] + " " + working["title_norm"]).str.strip()
    working["fuzzy_bucket"] = (
        working["company_norm"].str.slice(0, 10).fillna("")
        + "|"
        + working["title_norm"].str.split().str[0].fillna("")
    )
    return working


def aggregate_group(group: pd.DataFrame) -> pd.Series:
    """Aggregates a duplicate group into one merged job row."""
    group = group.copy()
    quality = (
        group["description_full"].fillna("").astype(str).str.len()
        + group["job_url"].fillna("").astype(str).str.len().clip(upper=1) * 200
        + group["salary_currency"].fillna("").astype(str).str.len().clip(upper=1) * 50
    )
    base_idx = quality.idxmax()
    base = group.loc[base_idx].copy()

    base["title"] = first_non_empty(group["title"]) or base.get("title")
    base["company"] = first_non_empty(group["company"]) or base.get("company")
    base["location"] = first_non_empty(group["location"]) or base.get("location")
    base["job_type"] = first_non_empty(group["job_type"]) or base.get("job_type")
    base["job_url"] = first_non_empty(group["job_url"]) or base.get("job_url")
    base["description_full"] = longest_text(group["description_full"])
    base["description_snippet"] = base["description_full"][:500]
    base["min_amount"] = first_non_empty(group["min_amount"])
    base["max_amount"] = first_non_empty(group["max_amount"])
    base["salary_currency"] = first_non_empty(group["salary_currency"])

    posted_dates = pd.to_datetime(group["date_posted"], errors="coerce")
    base["date_posted"] = posted_dates.max().date() if posted_dates.notna().any() else None

    base["found_on_platforms"] = unique_join(
        group["found_on_platforms"],
        split_pattern=r"\s*,\s*",
        separator=", ",
    )
    base["combo_that_found_it"] = unique_join(
        group["combo_that_found_it"],
        split_pattern=r"\s*\|\|\s*",
        separator=" || ",
    )
    base["search_location"] = unique_join(
        group["search_location"],
        split_pattern=r"\s*,\s*",
        separator=", ",
    )
    base["is_remote"] = bool(group["is_remote"].fillna(False).astype(bool).any())

    platforms = [item.strip() for item in str(base["found_on_platforms"]).split(",") if item.strip()]
    base["multi_platform_hit"] = len(platforms) >= 2

    return base


def merge_by_keys(df: pd.DataFrame, keys: list[str]) -> tuple[pd.DataFrame, int]:
    """Merges duplicates grouped by exact key columns."""
    if df.empty:
        return df.copy(), 0
    grouped_rows = [aggregate_group(group) for _, group in df.groupby(keys, dropna=False, sort=False)]
    merged = pd.DataFrame(grouped_rows).reset_index(drop=True)
    removed = max(0, len(df) - len(merged))
    return merged, removed


def fuzzy_deduplicate(df: pd.DataFrame, threshold: int = 92) -> tuple[pd.DataFrame, int]:
    """Performs conservative fuzzy dedup using RapidFuzz on company+title."""
    if df.empty:
        return df.copy(), 0

    merged_rows: list[pd.Series] = []
    removed = 0

    for _, bucket_df in df.groupby("fuzzy_bucket", dropna=False, sort=False):
        pending = list(bucket_df.index)
        while pending:
            anchor = pending.pop(0)
            anchor_key = str(df.at[anchor, "company_title_norm"])
            cluster = [anchor]
            remaining: list[int] = []

            for idx in pending:
                candidate_key = str(df.at[idx, "company_title_norm"])
                score = fuzz.token_sort_ratio(anchor_key, candidate_key)
                if score > threshold:
                    cluster.append(idx)
                else:
                    remaining.append(idx)

            pending = remaining
            cluster_df = df.loc[cluster]
            if len(cluster) > 1:
                removed += len(cluster) - 1
            merged_rows.append(aggregate_group(cluster_df))

    merged = pd.DataFrame(merged_rows).reset_index(drop=True)
    return merged, removed


def deduplicate_jobs(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int]]:
    """Runs the requested 4-step dedup pipeline and returns stats."""
    if df.empty:
        return df.copy(), {
            "step1_exact_removed": 0,
            "step2_url_removed": 0,
            "step3_fuzzy_removed": 0,
            "step4_cross_platform_removed": 0,
            "total_removed": 0,
        }

    step0 = add_normalized_keys(df)

    step1, removed1 = merge_by_keys(step0, ["company_norm", "title_norm", "location_norm", "date_norm"])
    step1 = add_normalized_keys(step1)

    with_url = step1[step1["job_url_norm"] != ""].copy()
    without_url = step1[step1["job_url_norm"] == ""].copy()

    if not with_url.empty:
        merged_url, removed2 = merge_by_keys(with_url, ["job_url_norm"])
        step2 = pd.concat([merged_url, without_url], ignore_index=True)
    else:
        step2 = step1.copy()
        removed2 = 0

    step2 = add_normalized_keys(step2)

    step3, removed3 = fuzzy_deduplicate(step2, threshold=92)
    step3 = add_normalized_keys(step3)

    step3["job_identity_key"] = step3["job_url_norm"]
    empty_identity = step3["job_identity_key"].eq("")
    step3.loc[empty_identity, "job_identity_key"] = (
        step3.loc[empty_identity, "company_norm"]
        + "|"
        + step3.loc[empty_identity, "title_norm"]
        + "|"
        + step3.loc[empty_identity, "location_norm"]
        + "|"
        + step3.loc[empty_identity, "date_norm"]
    )

    step4, removed4 = merge_by_keys(step3, ["job_identity_key"])

    cleanup_columns = [
        "company_norm",
        "title_norm",
        "location_norm",
        "date_norm",
        "job_url_norm",
        "company_title_norm",
        "fuzzy_bucket",
        "job_identity_key",
    ]
    final_df = step4.drop(columns=[col for col in cleanup_columns if col in step4.columns])
    final_df["found_on_platforms"] = final_df["found_on_platforms"].fillna("").astype(str).str.lower()
    final_df["multi_platform_hit"] = final_df["found_on_platforms"].apply(
        lambda value: len([item for item in str(value).split(",") if item.strip()]) >= 2
    )

    stats = {
        "step1_exact_removed": removed1,
        "step2_url_removed": removed2,
        "step3_fuzzy_removed": removed3,
        "step4_cross_platform_removed": removed4,
        "total_removed": removed1 + removed2 + removed3 + removed4,
    }
    return final_df.reset_index(drop=True), stats


def compile_skill_patterns(skills: list[str]) -> dict[str, re.Pattern[str]]:
    """Compiles skill regex patterns for faster matching."""
    patterns: dict[str, re.Pattern[str]] = {}
    for skill in skills:
        escaped = re.escape(skill)
        if re.fullmatch(r"[A-Za-z0-9]+", skill):
            pattern = re.compile(rf"\b{escaped}\b", re.IGNORECASE)
        else:
            pattern = re.compile(escaped, re.IGNORECASE)
        patterns[skill] = pattern
    return patterns


def extract_skills(text: str, skill_patterns: dict[str, re.Pattern[str]]) -> list[str]:
    """Extracts matching skills from text using the configured skill bank."""
    matches: list[str] = []
    for skill, pattern in skill_patterns.items():
        if pattern.search(text):
            matches.append(skill)
    return matches


def extract_experience_mentions(text: str) -> str:
    """Extracts concise experience phrases for reporting."""
    found: list[str] = []
    seen: set[str] = set()
    for match in EXPERIENCE_MENTION_PATTERN.finditer(text):
        token = " ".join(match.group(0).split())
        lower = token.lower()
        if lower not in seen:
            seen.add(lower)
            found.append(token)
    return ", ".join(found[:5])


def detect_bond_clause(description: str) -> tuple[bool, str]:
    """Detects bond-like clauses and returns the first relevant sentence."""
    if not description:
        return False, ""

    lower = description.lower()
    if not any(term in lower for term in BOND_TERMS):
        return False, ""

    sentences = re.split(r"(?<=[.!?])\s+|\n+", description)
    for sentence in sentences:
        sentence_clean = sentence.strip()
        lower_sentence = sentence_clean.lower()
        if any(term in lower_sentence for term in BOND_TERMS):
            return True, sentence_clean[:500]

    return True, description[:500]


def is_remote_job(row: pd.Series) -> bool:
    """Determines if a job should be treated as remote."""
    location_text = str(row.get("location", "") or "").lower()
    search_location = str(row.get("search_location", "") or "").lower()
    return bool(row.get("is_remote", False)) or "remote" in location_text or "remote" in search_location


def has_salary_signal(row: pd.Series) -> bool:
    """Checks if salary/CTC signal exists in structured data or text."""
    if pd.notna(row.get("min_amount")) or pd.notna(row.get("max_amount")):
        return True
    description = str(row.get("description_full", "") or "")
    return bool(re.search(r"(\u20b9|inr|ctc|lpa|salary|\$|usd|eur)", description, re.IGNORECASE))


def calculate_relevance_score(
    row: pd.Series,
    skills_found: list[str],
    today: date,
) -> int:
    """Calculates the requested relevance score without filtering."""
    score = 0
    title_text = str(row.get("title", "") or "")
    description_text = str(row.get("description_full", "") or "")

    if TITLE_FRESHER_SIGNAL_PATTERN.search(title_text):
        score += 25
    if DESC_POSITIVE_EXPERIENCE_PATTERN.search(description_text):
        score += 20
    if skills_found:
        score += 15

    posted = pd.to_datetime(row.get("date_posted"), errors="coerce")
    if pd.notna(posted) and (today - posted.date()).days <= 7:
        score += 10

    if bool(row.get("multi_platform_hit", False)):
        score += 10
    if is_remote_job(row):
        score += 10
    if CORE_CS_PATTERN.search(description_text):
        score += 5
    if has_salary_signal(row):
        score += 5
    if DESC_THREE_PLUS_PATTERN.search(description_text):
        score -= 10
    if DESC_FIVE_PLUS_PATTERN.search(description_text):
        score -= 20

    return score


def enrich_jobs(df: pd.DataFrame, skills_bank: list[str]) -> pd.DataFrame:
    """Adds scoring, skills, bond detection, and reporting columns."""
    if df.empty:
        return df

    working = df.copy()
    skill_patterns = compile_skill_patterns(skills_bank)
    today = datetime.now().date()

    skills_col: list[str] = []
    exp_col: list[str] = []
    bond_flag_col: list[bool] = []
    bond_details_col: list[str] = []
    relevance_col: list[int] = []
    remote_col: list[bool] = []

    for _, row in working.iterrows():
        combined_text = f"{row.get('title', '')} {row.get('description_full', '')}"
        skills_found = extract_skills(combined_text, skill_patterns)
        experience_mentioned = extract_experience_mentions(combined_text)
        bond_flag, bond_details = detect_bond_clause(str(row.get("description_full", "") or ""))
        relevance = calculate_relevance_score(row, skills_found, today)
        remote_flag = is_remote_job(row)

        skills_col.append(", ".join(skills_found))
        exp_col.append(experience_mentioned)
        bond_flag_col.append(bond_flag)
        bond_details_col.append(bond_details)
        relevance_col.append(int(relevance))
        remote_col.append(remote_flag)

    working["skills_matched"] = skills_col
    working["experience_mentioned"] = exp_col
    working["bond_flag"] = bond_flag_col
    working["bond_details"] = bond_details_col
    working["relevance_score"] = relevance_col
    working["is_remote"] = remote_col
    working["description_snippet"] = working["description_full"].fillna("").astype(str).str.slice(0, 500)
    working["date_posted"] = pd.to_datetime(working["date_posted"], errors="coerce").dt.date

    for col in OUTPUT_COLUMNS:
        if col not in working.columns:
            working[col] = None

    remaining_columns = [col for col in working.columns if col not in OUTPUT_COLUMNS]
    working = working[OUTPUT_COLUMNS + remaining_columns]
    working = working.sort_values(by=["relevance_score", "date_posted"], ascending=[False, False]).reset_index(drop=True)
    return working


def score_fill(score: Any) -> PatternFill:
    """Returns the worksheet row fill based on relevance score."""
    try:
        value = float(score)
    except (TypeError, ValueError):
        value = 0.0
    if value > 80:
        return COLOR_DARK_GREEN
    if value > 60:
        return COLOR_LIGHT_GREEN
    if value > 40:
        return COLOR_YELLOW
    if value > 20:
        return COLOR_ORANGE
    return COLOR_RED


def auto_fit_columns(worksheet, max_width: int = 80) -> None:
    """Auto-fits worksheet column widths with a max cap."""
    for column_cells in worksheet.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_len + 2, max_width)


def format_all_jobs_sheet(worksheet) -> None:
    """Applies color scale, hyperlinking, filters, and freeze pane to All Jobs."""
    headers = {cell.value: idx for idx, cell in enumerate(worksheet[1], start=1)}
    relevance_col = headers.get("relevance_score")
    company_col = headers.get("company")
    bond_col = headers.get("bond_flag")
    url_col = headers.get("job_url")

    if relevance_col is None:
        return

    for row_idx in range(2, worksheet.max_row + 1):
        fill = score_fill(worksheet.cell(row=row_idx, column=relevance_col).value)
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = fill

        if bond_col and company_col:
            bond_value = str(worksheet.cell(row=row_idx, column=bond_col).value).strip().lower()
            if bond_value in {"true", "1", "yes"}:
                worksheet.cell(row=row_idx, column=company_col).font = BOND_FONT

        if url_col:
            url_cell = worksheet.cell(row=row_idx, column=url_col)
            url_value = str(url_cell.value or "").strip()
            if url_value:
                url_cell.hyperlink = url_value
                url_cell.style = "Hyperlink"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    auto_fit_columns(worksheet)


def highlight_bond_sheet(worksheet) -> None:
    """Highlights bond details in the Bond Flagged sheet."""
    headers = {cell.value: idx for idx, cell in enumerate(worksheet[1], start=1)}
    bond_details_col = headers.get("bond_details")
    if bond_details_col is None:
        return

    highlight = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_idx, column=bond_details_col).fill = highlight

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    auto_fit_columns(worksheet)


def build_summary_tables(
    final_df: pd.DataFrame,
    raw_count: int,
    dedup_stats: dict[str, int],
) -> list[tuple[str, pd.DataFrame]]:
    """Builds summary dashboard tables for the Excel report."""
    summary_tables: list[tuple[str, pd.DataFrame]] = []

    summary_tables.append(
        (
            "Run Summary",
            pd.DataFrame(
                {
                    "metric": [
                        "Total jobs found (before dedup)",
                        "Total jobs found (after dedup)",
                        "Step 1 exact duplicates removed",
                        "Step 2 URL duplicates removed",
                        "Step 3 fuzzy duplicates removed",
                        "Step 4 cross-platform merges",
                        "Total duplicates removed",
                        "Bond jobs flagged",
                    ],
                    "value": [
                        raw_count,
                        len(final_df),
                        dedup_stats.get("step1_exact_removed", 0),
                        dedup_stats.get("step2_url_removed", 0),
                        dedup_stats.get("step3_fuzzy_removed", 0),
                        dedup_stats.get("step4_cross_platform_removed", 0),
                        dedup_stats.get("total_removed", 0),
                        int(final_df.get("bond_flag", pd.Series(dtype=bool)).fillna(False).sum()),
                    ],
                }
            ),
        )
    )

    platform_counts = (
        final_df.assign(platform=final_df["found_on_platforms"].fillna("").astype(str).str.split(","))
        .explode("platform")
    )
    platform_counts["platform"] = platform_counts["platform"].fillna("").astype(str).str.strip()
    platform_counts = platform_counts[platform_counts["platform"] != ""]
    platform_breakdown = (
        platform_counts["platform"].value_counts().rename_axis("platform").reset_index(name="job_count")
    )
    summary_tables.append(("Breakdown by Platform", platform_breakdown))

    location_breakdown = (
        final_df["location"].fillna("Unknown").astype(str).value_counts().head(50)
        .rename_axis("location")
        .reset_index(name="job_count")
    )
    summary_tables.append(("Breakdown by Location", location_breakdown))

    company_breakdown = (
        final_df["company"].fillna("Unknown").astype(str).value_counts().head(30)
        .rename_axis("company")
        .reset_index(name="job_count")
    )
    summary_tables.append(("Top 30 Companies", company_breakdown))

    skill_breakdown = (
        final_df.assign(skill=final_df["skills_matched"].fillna("").astype(str).str.split(","))
        .explode("skill")
    )
    skill_breakdown["skill"] = skill_breakdown["skill"].fillna("").astype(str).str.strip()
    skill_breakdown = skill_breakdown[skill_breakdown["skill"] != ""]
    top_skills = skill_breakdown["skill"].value_counts().head(20).rename_axis("skill").reset_index(name="count")
    summary_tables.append(("Top 20 Skills Mentioned", top_skills))

    title_breakdown = (
        final_df["title"].fillna("Unknown").astype(str).value_counts().head(10)
        .rename_axis("title")
        .reset_index(name="job_count")
    )
    summary_tables.append(("Top 10 Job Titles", title_breakdown))

    today = datetime.now().date()

    def date_bucket(posted: Any) -> str:
        value = pd.to_datetime(posted, errors="coerce")
        if pd.isna(value):
            return "Unknown"
        delta = (today - value.date()).days
        if delta <= 7:
            return "Last 7 days"
        if delta <= 15:
            return "8-15 days"
        if delta <= 30:
            return "16-30 days"
        return "Older than 30 days"

    date_breakdown = (
        final_df["date_posted"].apply(date_bucket).value_counts()
        .rename_axis("bucket")
        .reset_index(name="job_count")
    )
    summary_tables.append(("Jobs by Date Posted", date_breakdown))

    remote_breakdown = (
        final_df["is_remote"].fillna(False).astype(bool).map({True: "Remote", False: "On-site/Hybrid"})
        .value_counts()
        .rename_axis("mode")
        .reset_index(name="job_count")
    )
    summary_tables.append(("Remote vs On-site", remote_breakdown))

    return summary_tables


def save_csv(df: pd.DataFrame, path: Path) -> None:
    """Saves DataFrame to CSV with safe quoting."""
    df.to_csv(path, index=False, quoting=csv.QUOTE_NONNUMERIC, escapechar="\\")


def save_json(df: pd.DataFrame, path: Path) -> None:
    """Saves DataFrame to JSON records for dashboard/API use."""
    records = df.to_dict(orient="records")

    def default_serializer(value: Any) -> str | Any:
        if isinstance(value, (datetime, date, pd.Timestamp)):
            return value.isoformat()
        return value

    with path.open("w", encoding="utf-8") as handle:
        json.dump(records, handle, indent=2, ensure_ascii=False, default=default_serializer)


def save_excel(
    all_jobs_df: pd.DataFrame,
    path: Path,
    raw_count: int,
    dedup_stats: dict[str, int],
) -> None:
    """Saves the required multi-sheet Excel workbook with formatting."""
    remote_df = all_jobs_df[all_jobs_df["is_remote"].fillna(False).astype(bool)].copy()
    remote_df = remote_df.sort_values(by=["relevance_score", "date_posted"], ascending=[False, False])

    bond_df = all_jobs_df[all_jobs_df["bond_flag"].fillna(False).astype(bool)].copy()
    high_conf_df = all_jobs_df[all_jobs_df["multi_platform_hit"].fillna(False).astype(bool)].copy()

    summary_tables = build_summary_tables(all_jobs_df, raw_count=raw_count, dedup_stats=dedup_stats)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        all_jobs_df.to_excel(writer, sheet_name="All Jobs", index=False)

        start_row = 0
        for title, table in summary_tables:
            pd.DataFrame([[title]]).to_excel(
                writer,
                sheet_name="Summary Dashboard",
                index=False,
                header=False,
                startrow=start_row,
            )
            start_row += 1
            if table.empty:
                table = pd.DataFrame({"info": ["No data"]})
            table.to_excel(
                writer,
                sheet_name="Summary Dashboard",
                index=False,
                startrow=start_row,
            )
            start_row += len(table) + 2

        remote_df.to_excel(writer, sheet_name="Remote Only", index=False)
        bond_df.to_excel(writer, sheet_name="Bond Flagged", index=False)
        high_conf_df.to_excel(writer, sheet_name="High Confidence", index=False)

        workbook = writer.book
        all_jobs_sheet = workbook["All Jobs"]
        summary_sheet = workbook["Summary Dashboard"]
        remote_sheet = workbook["Remote Only"]
        bond_sheet = workbook["Bond Flagged"]
        high_conf_sheet = workbook["High Confidence"]

        format_all_jobs_sheet(all_jobs_sheet)
        summary_sheet.freeze_panes = "A2"
        summary_sheet.auto_filter.ref = summary_sheet.dimensions
        auto_fit_columns(summary_sheet)

        for sheet in [remote_sheet, high_conf_sheet]:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            auto_fit_columns(sheet)

        highlight_bond_sheet(bond_sheet)


def load_seen_urls(path: Path) -> set[str]:
    """Loads historical seen job URLs from disk."""
    if not path.exists():
        return set()
    try:
        seen_df = pd.read_csv(path)
    except Exception:
        return set()
    if "job_url" not in seen_df.columns:
        return set()
    return {canonicalize_url(url) for url in seen_df["job_url"].dropna().astype(str)}


def update_seen_urls(path: Path, new_urls: list[str], timestamp: str) -> None:
    """Appends new URLs to seen_jobs.csv with first-seen timestamp."""
    if not new_urls:
        return

    unique_new_urls = sorted({canonicalize_url(url) for url in new_urls if canonicalize_url(url)})
    if not unique_new_urls:
        return

    records_df = pd.DataFrame({"job_url": unique_new_urls, "first_seen_at": timestamp})
    if path.exists():
        try:
            existing = pd.read_csv(path)
            if "job_url" in existing.columns:
                existing_urls = {canonicalize_url(url) for url in existing["job_url"].dropna().astype(str)}
                records_df = records_df[~records_df["job_url"].isin(existing_urls)]
        except Exception:
            pass

    if records_df.empty:
        return

    if path.exists():
        records_df.to_csv(path, mode="a", index=False, header=False)
    else:
        records_df.to_csv(path, index=False)


def apply_incremental_mode(
    df: pd.DataFrame,
    seen_jobs_path: Path,
    timestamp: str,
) -> tuple[pd.DataFrame, int, int]:
    """Returns new jobs only and updates seen_jobs history."""
    if df.empty:
        return df.copy(), 0, 0

    seen_urls = load_seen_urls(seen_jobs_path)
    job_url_norm = df["job_url"].apply(canonicalize_url)

    duplicate_mask = job_url_norm.apply(lambda url: bool(url) and url in seen_urls)
    new_jobs_df = df.loc[~duplicate_mask].copy()
    history_duplicates = int(duplicate_mask.sum())

    new_urls = job_url_norm.loc[~duplicate_mask].tolist()
    update_seen_urls(seen_jobs_path, new_urls, timestamp)

    return new_jobs_df, history_duplicates, len([url for url in new_urls if url])


def parse_arguments() -> argparse.Namespace:
    """Defines and parses CLI arguments."""
    parser = argparse.ArgumentParser(description="Enhanced JobSpy fresher discovery scraper")
    parser.add_argument("--titles", type=str, default=None, help="Comma-separated job titles")
    parser.add_argument("--skills", type=str, default=None, help="Comma-separated skills")
    parser.add_argument("--locations", type=str, default=None, help="Comma-separated locations")
    parser.add_argument("--days", type=int, default=None, help="Max age of jobs in days")
    parser.add_argument("--results", type=int, default=None, help="Results per combo per platform")
    parser.add_argument("--platforms", type=str, default=None, help="Comma-separated platforms")
    parser.add_argument(
        "--output",
        type=str,
        choices=["csv", "excel", "json", "all"],
        default="all",
        help="Output format",
    )
    parser.add_argument(
        "--new-only",
        type=parse_bool,
        default=True,
        help="Save incremental NEW jobs file (True/False)",
    )
    parser.add_argument("--proxies", type=str, default=None, help="Path to proxies.txt")
    parser.add_argument("--workers", type=int, default=None, help="Parallel worker count")
    parser.add_argument("--checkpoint", type=int, default=None, help="Checkpoint interval")
    parser.add_argument("--config", type=str, default=str(DEFAULT_CONFIG_PATH), help="Path to config.yaml")
    return parser.parse_args()


def resolve_runtime_settings(args: argparse.Namespace, config: dict[str, Any]) -> dict[str, Any]:
    """Resolves runtime settings from CLI with config defaults."""
    defaults = config.get("defaults", {})

    min_combos = int(defaults.get("min_combos", 100))
    days = int(args.days if args.days is not None else defaults.get("days", 30))
    results = int(args.results if args.results is not None else defaults.get("results", 50))
    workers = int(args.workers if args.workers is not None else defaults.get("workers", 3))
    checkpoint_every = int(
        args.checkpoint if args.checkpoint is not None else defaults.get("checkpoint", 25)
    )

    retry_backoff = defaults.get("retry_backoff_seconds", [10, 30, 60])
    if not isinstance(retry_backoff, list) or len(retry_backoff) < 3:
        retry_backoff = [10, 30, 60]

    sleep_range = defaults.get("random_sleep_range_seconds", [3, 8])
    if (
        not isinstance(sleep_range, list)
        or len(sleep_range) < 2
        or float(sleep_range[0]) > float(sleep_range[1])
    ):
        sleep_range = [3, 8]

    return {
        "min_combos": max(100, min_combos),
        "days": max(1, days),
        "results": max(1, results),
        "workers": max(1, workers),
        "checkpoint_every": max(1, checkpoint_every),
        "retry_backoff": [int(retry_backoff[0]), int(retry_backoff[1]), int(retry_backoff[2])],
        "sleep_range": (float(sleep_range[0]), float(sleep_range[1])),
        "country_indeed": str(defaults.get("country_indeed", "India")),
        "seen_jobs_file": Path(str(defaults.get("seen_jobs_file", DEFAULT_SEEN_JOBS_PATH))),
    }


def save_outputs(
    final_df: pd.DataFrame,
    timestamp: str,
    output_mode: str,
    raw_count: int,
    dedup_stats: dict[str, int],
) -> dict[str, Path]:
    """Saves requested all-jobs output files and returns their paths."""
    output_paths: dict[str, Path] = {}

    if output_mode in {"csv", "all"}:
        csv_path = Path(f"jobs_ALL_{timestamp}.csv")
        save_csv(final_df, csv_path)
        output_paths["csv"] = csv_path

    if output_mode in {"excel", "all"}:
        xlsx_path = Path(f"jobs_ALL_{timestamp}.xlsx")
        save_excel(final_df, xlsx_path, raw_count=raw_count, dedup_stats=dedup_stats)
        output_paths["excel"] = xlsx_path

    if output_mode in {"json", "all"}:
        json_path = Path(f"jobs_ALL_{timestamp}.json")
        save_json(final_df, json_path)
        output_paths["json"] = json_path

    return output_paths


def save_new_jobs_outputs(new_jobs_df: pd.DataFrame, timestamp: str) -> dict[str, Path]:
    """Saves incremental new jobs files (CSV + Excel)."""
    output_paths: dict[str, Path] = {}

    csv_path = Path(f"new_jobs_{timestamp}.csv")
    save_csv(new_jobs_df, csv_path)
    output_paths["csv"] = csv_path

    xlsx_path = Path(f"new_jobs_{timestamp}.xlsx")
    save_excel(
        new_jobs_df,
        xlsx_path,
        raw_count=len(new_jobs_df),
        dedup_stats={
            "step1_exact_removed": 0,
            "step2_url_removed": 0,
            "step3_fuzzy_removed": 0,
            "step4_cross_platform_removed": 0,
            "total_removed": 0,
        },
    )
    output_paths["excel"] = xlsx_path

    return output_paths


def main() -> int:
    """Entry point for the enhanced fresher job scraper."""
    args = parse_arguments()
    run_started = datetime.now()
    timestamp = run_started.strftime("%Y%m%d_%H%M%S")
    logger, log_path = setup_logging(timestamp)

    try:
        config = load_config(Path(args.config))

        job_titles = resolve_bank(config, "job_titles", args.titles)
        skills = resolve_bank(config, "skills", args.skills)
        experience_qualifiers = resolve_bank(config, "experience_qualifiers", None)
        locations = resolve_bank(config, "locations", args.locations)
        platforms = resolve_platforms(config, args.platforms)
        proxies = load_proxies(args.proxies)

        settings = resolve_runtime_settings(args, config)

        combos = generate_search_combinations(
            job_titles=job_titles,
            skills=skills,
            experience_qualifiers=experience_qualifiers,
            locations=locations,
            min_combos=settings["min_combos"],
        )

        tasks = build_combo_tasks(combos, locations)
        logger.info("Generated %s unique keyword combinations", len(tasks))
        logger.info("Platforms: %s", ", ".join(platforms))
        logger.info("Results per combo per platform: %s", settings["results"])
        logger.info("Jobs age window: %s days", settings["days"])

        user_agent_getter = initialize_user_agent_provider(logger)
        hours_old = settings["days"] * 24

        raw_frames: list[pd.DataFrame] = []
        completed = 0

        with ThreadPoolExecutor(max_workers=settings["workers"]) as executor:
            futures = {
                executor.submit(
                    run_combo_task,
                    task,
                    platforms=platforms,
                    results_wanted=settings["results"],
                    hours_old=hours_old,
                    country_indeed=settings["country_indeed"],
                    proxies=proxies,
                    user_agent_getter=user_agent_getter,
                    retry_backoff_seconds=settings["retry_backoff"],
                    sleep_range_seconds=settings["sleep_range"],
                    scrape_timestamp=timestamp,
                    logger=logger,
                ): task
                for task in tasks
            }

            with tqdm(total=len(tasks), desc="Scraping combos", unit="combo") as progress:
                for future in as_completed(futures):
                    result = future.result()
                    completed += 1
                    progress.update(1)

                    if isinstance(result.jobs_df, pd.DataFrame) and not result.jobs_df.empty:
                        raw_frames.append(result.jobs_df)

                    task = result.task
                    platform_counts = result.platform_counts
                    platform_msg = " | ".join(
                        [
                            f"{platform.title().replace('_', ' ')}: {platform_counts.get(platform, 0)}"
                            for platform in platforms
                        ]
                    )
                    logger.info(
                        "Combo %s/%s | '%s' | %s | Total: %s",
                        task.order,
                        len(tasks),
                        task.combo.query,
                        platform_msg,
                        sum(platform_counts.values()),
                    )

                    if result.skipped_platforms:
                        logger.warning(
                            "Combo %s skipped platforms: %s",
                            task.combo.combo_id,
                            ", ".join(result.skipped_platforms),
                        )

                    if completed % settings["checkpoint_every"] == 0:
                        checkpoint_path = save_checkpoint(raw_frames, timestamp, completed)
                        if checkpoint_path:
                            logger.info("Checkpoint saved: %s", checkpoint_path)

        if raw_frames:
            raw_df = pd.concat(raw_frames, ignore_index=True)
        else:
            raw_df = pd.DataFrame()

        standardized_df = standardize_raw_jobs(raw_df) if not raw_df.empty else pd.DataFrame(columns=OUTPUT_COLUMNS)
        total_raw_results = len(standardized_df)

        filtered_df = apply_fresher_filter(standardized_df) if not standardized_df.empty else standardized_df
        deduped_df, dedup_stats = deduplicate_jobs(filtered_df)
        final_df = enrich_jobs(deduped_df, skills_bank=skills)

        if final_df.empty:
            logger.warning("No jobs found after filtering and deduplication.")

        all_output_paths = save_outputs(
            final_df,
            timestamp=timestamp,
            output_mode=args.output,
            raw_count=total_raw_results,
            dedup_stats=dedup_stats,
        )

        if args.new_only:
            new_jobs_df, history_dupes, new_url_count = apply_incremental_mode(
                final_df,
                seen_jobs_path=settings["seen_jobs_file"],
                timestamp=timestamp,
            )
            new_output_paths = save_new_jobs_outputs(new_jobs_df, timestamp)
        else:
            history_dupes = 0
            new_url_count = 0
            new_jobs_df = pd.DataFrame()
            new_output_paths = {}

        bond_jobs_count = int(final_df.get("bond_flag", pd.Series(dtype=bool)).fillna(False).sum())

        logger.info("Run complete")
        logger.info("Total combos run: %s", len(tasks))
        logger.info("Total raw results: %s", total_raw_results)
        logger.info("Duplicates removed: %s", dedup_stats.get("total_removed", 0))
        logger.info("Bond jobs flagged: %s", bond_jobs_count)
        logger.info("Final job count: %s", len(final_df))
        logger.info("New jobs (vs history): %s", len(new_jobs_df))
        logger.info("Historical duplicates skipped: %s", history_dupes)
        logger.info("New job URLs appended to history: %s", new_url_count)
        logger.info("Log file: %s", log_path)

        for label, path in all_output_paths.items():
            logger.info("Saved %s output: %s", label.upper(), path)
        for label, path in new_output_paths.items():
            logger.info("Saved NEW-JOBS %s output: %s", label.upper(), path)

        return 0

    except Exception as exc:
        logger.exception("Run failed: %s", exc)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
